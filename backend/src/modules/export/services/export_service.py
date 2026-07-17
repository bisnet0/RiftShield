from __future__ import annotations

import csv
import io
import json
import zipfile
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional

BR_TZ = timezone(timedelta(hours=-3))


def _now_br():
    return datetime.now(timezone.utc).astimezone(BR_TZ)


async def _get_inferences(user_id: str) -> List[Dict]:
    from modules.inference.models.inference_model import InferenceResult
    items = await InferenceResult.find({"user_id": user_id}).sort(-InferenceResult.created_at).limit(1000).to_list()
    return [
        {"id": str(i.id), "filename": i.filename, "status": i.status,
         "components": [{"label": c.label if hasattr(c, "model_dump") else c.get("label"), "confidence": c.confidence if hasattr(c, "model_dump") else c.get("confidence")} for c in (i.components or [])],
         "processing_time_ms": i.processing_time_ms, "fallback_used": i.fallback_used,
         "created_at": i.created_at.isoformat() if i.created_at else None}
        for i in items
    ]


async def _get_threats(user_id: str) -> List[Dict]:
    from modules.inference.models.threat_model import ThreatReport
    items = await ThreatReport.find({"user_id": user_id}).sort(-ThreatReport.created_at).limit(1000).to_list()
    return [
        {"id": str(r.id), "inference_id": r.inference_id, "status": r.status,
         "stride_summary": r.stride_summary, "overall_risk_score": r.overall_risk_score,
         "created_at": r.created_at.isoformat() if r.created_at else None}
        for r in items
    ]


async def _get_dataset(user_id: str) -> List[Dict]:
    from modules.inference.dataset.dataset_model import DatasetEntry
    items = await DatasetEntry.find({"user_id": user_id}).sort(-DatasetEntry.created_at).limit(1000).to_list()
    return [
        {"id": str(e.id), "filename": e.filename, "split": e.split, "source": e.source,
         "labels": [{"label": l.label, "class_id": l.class_id} for l in (e.labels or [])],
         "created_at": e.created_at.isoformat() if e.created_at else None}
        for e in items
    ]


async def _get_training_logs(user_id: str) -> List[Dict]:
    from modules.inference.models.inference_model import TrainingLog
    items = await TrainingLog.find({}).sort(-TrainingLog.created_at).limit(100).to_list()
    return [
        {"id": str(t.id), "model_type": t.model_type, "status": t.status,
         "hyperparameters": t.hyperparameters, "metrics": t.metrics,
         "created_at": t.created_at.isoformat() if t.created_at else None}
        for t in items
    ]


def _pick(obj: dict, lang: str, field: str) -> str:
    key = f"{field}_{lang[:2].lower()}"
    return obj.get(key) or obj.get(f"{field}_pt") or ""


async def _get_vulnerabilities(lang: str = "pt-BR") -> List[Dict]:
    from modules.inference.models.kb_model import KBVulnerability
    items = await KBVulnerability.find({}).sort(-KBVulnerability.created_at).limit(500).to_list()
    return [
        {"cve_id": v.cve_id, "title": _pick(v.model_dump(), lang, "title"), "description": _pick(v.model_dump(), lang, "description"),
         "cvss_score": v.cvss_score, "cwe": v.cwe,
         "affected_components": v.affected_components, "tags": v.tags}
        for v in items
    ]


async def _get_countermeasures(lang: str = "pt-BR") -> List[Dict]:
    from modules.inference.models.kb_model import KBCountermeasure
    items = await KBCountermeasure.find({}).limit(500).to_list()
    return [
        {"title": _pick(c.model_dump(), lang, "title"), "description": _pick(c.model_dump(), lang, "description"),
         "priority": c.priority,
         "implementation_guide": c.implementation_guide, "references": c.references,
         "vulnerability_cwe_ids": c.vulnerability_cwe_ids}
        for c in items
    ]


async def _get_profile(user_id: str) -> Optional[Dict]:
    from modules.auth.models.user_model import User
    user = await User.get(user_id)
    if not user:
        return None
    return {"name": user.name, "email": user.email, "phone": user.phone,
            "country": user.country, "state": user.state, "city": user.city,
            "role": user.role, "profession": user.profession, "seniority": user.seniority,
            "age": user.age, "total_days_active": user.total_days_active, "language": user.language}


async def _get_settings(user_id: str) -> Optional[Dict]:
    from modules.hermes.models.llm_config import HermesConfig
    config = await HermesConfig.find_one(HermesConfig.user_id == user_id)
    if not config:
        return None
    return {"enabled": config.enabled, "provider": config.provider, "diag_fallback": config.diag_fallback}


async def _get_comparisons(user_id: str) -> List[Dict]:
    from modules.inference.models.comparison_model import ComparisonLog
    items = await ComparisonLog.find({"user_id": user_id}).sort(-ComparisonLog.created_at).limit(50).to_list()
    return [
        {
            "id": str(i.id),
            "filename_a": i.filename_a,
            "filename_b": i.filename_b,
            "verdict": i.result.get("verdict", ""),
            "risk_delta": i.result.get("diff", {}).get("risk_delta"),
            "has_suggestion": i.suggestion is not None,
            "created_at": i.created_at.isoformat() if i.created_at else None,
        }
        for i in items
    ]


async def export_data(
    user_id: str,
    sections: List[str],
    include_profile: bool = False,
    include_settings: bool = False,
    fmt: str = "json",
    zip_output: bool = False,
    lang: str = "pt-BR",
) -> Any:
    data: Dict[str, Any] = {}

    if "inferences" in sections:
        data["analise_de_diagramas"] = await _get_inferences(user_id)
    if "threats" in sections:
        data["relatorios_stride"] = await _get_threats(user_id)
    if "dataset" in sections:
        data["dataset"] = await _get_dataset(user_id)
    if "training" in sections:
        data["treinamento"] = await _get_training_logs(user_id)
    if "vulnerabilities" in sections:
        data["vulnerabilidades"] = await _get_vulnerabilities(lang)
    if "countermeasures" in sections:
        data["contramedidas"] = await _get_countermeasures(lang)
    if "comparisons" in sections:
        data["comparacoes"] = await _get_comparisons(user_id)
    if include_profile:
        profile = await _get_profile(user_id)
        if profile:
            data["perfil"] = profile
    if include_settings:
        settings = await _get_settings(user_id)
        if settings:
            data["configuracoes"] = settings

    if fmt == "pdf":
        from modules.export.services.pdf_generator import build_pdf_sections, generate_pdf
        sections_data = build_pdf_sections(
            inferences=data.get("analise_de_diagramas", []),
            threats=data.get("relatorios_stride", []),
            dataset=data.get("dataset", []),
            training_logs=data.get("treinamento", []),
            vulnerabilities=data.get("vulnerabilidades", []),
            countermeasures=data.get("contramedidas", []),
            profile=data.get("perfil"),
            settings=data.get("configuracoes"),
            lang=lang,
        )
        now = _now_br()
        fmt_date = "%d/%m/%Y às %H:%M" if lang.startswith("pt") else "%m/%d/%Y at %H:%M"
        generation_date = now.strftime(fmt_date)
        pdf_bytes = generate_pdf(sections_data, generation_date, data.get("perfil"), lang)
        filename = f"riftshield_relatorio_{now.strftime('%Y%m%d_%H%M%S')}.pdf"
        return await _maybe_zip(pdf_bytes, filename, zip_output)

    if fmt == "json":
        content = json.dumps(data, ensure_ascii=False, indent=2, default=str).encode("utf-8")
        filename = f"riftshield_export_{_now_br().strftime('%Y%m%d_%H%M%S')}.json"
        return await _maybe_zip(content, filename, zip_output)

    elif fmt == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        for section, items in data.items():
            writer.writerow([f"=== {section} ==="])
            if items and isinstance(items, list) and len(items) > 0:
                writer.writerow(list(items[0].keys()))
                for item in items:
                    writer.writerow([str(v) for v in item.values()])
            writer.writerow([])
        content = buffer.getvalue().encode("utf-8")
        filename = f"riftshield_export_{_now_br().strftime('%Y%m%d_%H%M%S')}.csv"
        return await _maybe_zip(content, filename, zip_output)

    elif fmt == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="E65C00", end_color="E65C00", fill_type="solid")

            def _flatten(val):
                if isinstance(val, list):
                    if not val:
                        return ""
                    if isinstance(val[0], dict):
                        return "; ".join(
                            f"{v.get('label','?')}({v.get('confidence',0):.2f})"
                            for v in val
                        )
                    return "; ".join(str(v) for v in val)
                if isinstance(val, dict):
                    return "; ".join(f"{k}={v}" for k, v in val.items())
                return str(val) if val is not None else ""

            for section, items in data.items():
                ws = wb.create_sheet(title=section[:31])
                if items and isinstance(items, list) and len(items) > 0:
                    first = items[0]
                    if isinstance(first, dict):
                        keys = list(first.keys())
                        for ci, k in enumerate(keys, 1):
                            cell = ws.cell(row=1, column=ci, value=k)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center", wrap_text=True)
                        for ri, item in enumerate(items, 2):
                            for ci, k in enumerate(keys, 1):
                                ws.cell(row=ri, column=ci, value=_flatten(item.get(k)))
                    elif isinstance(first, (str, int, float)):
                        for ri, item in enumerate(items, 1):
                            ws.cell(row=ri, column=1, value=str(item))
                for col_cells in ws.columns:
                    vals = [str(c.value or "") for c in col_cells]
                    max_len = max(len(v) for v in vals) if vals else 10
                    ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 80)
            wb.remove(wb.active)
            buffer = io.BytesIO()
            wb.save(buffer)
            content = buffer.getvalue()
            filename = f"riftshield_export_{_now_br().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return await _maybe_zip(content, filename, zip_output)
        except ImportError:
            return {"error": "openpyxl not installed. Use JSON or CSV instead."}

    return {"error": f"Formato '{fmt}' não suportado"}


async def _maybe_zip(content: bytes, filename: str, zip_output: bool) -> Any:
    if not zip_output:
        if isinstance(content, bytes):
            try:
                decoded = content.decode("utf-8")
                return {"filename": filename, "content": decoded}
            except UnicodeDecodeError:
                return {"filename": filename, "content": list(content)}
        return {"filename": filename, "content": content}

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(filename, content if isinstance(content, bytes) else content.encode("utf-8") if isinstance(content, str) else content)
    zip_name = filename.rsplit(".", 1)[0] + ".zip"
    return {"filename": zip_name, "content": list(buffer.getvalue())}
